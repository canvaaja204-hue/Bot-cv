# main.py
import telebot
from telebot import types
import os
import time
import csv
from openpyxl import load_workbook

# ===== CONFIG =====
TOKEN = "8195532623:AAF-87A-2Bkq8T9JAEjKsHB4Dzxfw7FltvQ"  # Ganti dengan token BotFather
ADMIN_ID = [6400667666]  # Ganti dengan Telegram ID kamu

bot = telebot.TeleBot(TOKEN)

# ===== HELPERS =====
def txt_to_vcf(txt_path, output_path):
    contacts = []
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            name, number = line.strip().split(',')
            contacts.append((name, number))
    with open(output_path, 'w', encoding='utf-8') as f:
        for name, number in contacts:
            f.write("BEGIN:VCARD\n")
            f.write("VERSION:3.0\n")
            f.write(f"N:{name}\n")
            f.write(f"TEL:{number}\n")
            f.write("END:VCARD\n")

def vcf_to_txt(vcf_path, output_path):
    contacts = []
    with open(vcf_path, 'r', encoding='utf-8') as f:
        name = number = ""
        for line in f:
            if line.startswith("N:"):
                name = line.strip()[2:]
            if line.startswith("TEL:"):
                number = line.strip()[4:]
            if line.strip() == "END:VCARD":
                contacts.append((name, number))
    with open(output_path, 'w', encoding='utf-8') as f:
        for name, number in contacts:
            f.write(f"{name},{number}\n")

def xlsx_to_vcf(xlsx_path, output_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    contacts = []
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1]:
            contacts.append((row[0], row[1]))
    with open(output_path, 'w', encoding='utf-8') as f:
        for name, number in contacts:
            f.write("BEGIN:VCARD\n")
            f.write("VERSION:3.0\n")
            f.write(f"N:{name}\n")
            f.write(f"TEL:{number}\n")
            f.write("END:VCARD\n")

# ===== START COMMAND =====
@bot.message_handler(commands=['start'])
def start_message(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.KeyboardButton("TXT → VCF")
    btn2 = types.KeyboardButton("VCF → TXT")
    btn3 = types.KeyboardButton("XLSX → VCF")
    btn4 = types.KeyboardButton("Hitung Kontak")
    btn5 = types.KeyboardButton("Rename Kontak")
    btn6 = types.KeyboardButton("Hapus Kontak")
    btn7 = types.KeyboardButton("Pecah File")
    btn8 = types.KeyboardButton("Gabung File")
    markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8)
    bot.send_message(message.chat.id, "Halo bro! Bot kontak siap 😎", reply_markup=markup)

# ===== HANDLE MENU =====
@bot.message_handler(func=lambda message: True)
def menu_handler(message):
    chat_id = message.chat.id
    text = message.text

    try:
        if text == "TXT → VCF":
            bot.send_message(chat_id, "Silakan kirim file TXT kamu dengan format: Nama,Nomor")
            bot.register_next_step_handler_by_chat_id(chat_id, handle_txt_to_vcf)
        elif text == "VCF → TXT":
            bot.send_message(chat_id, "Silakan kirim file VCF kamu")
            bot.register_next_step_handler_by_chat_id(chat_id, handle_vcf_to_txt)
        elif text == "XLSX → VCF":
            bot.send_message(chat_id, "Silakan kirim file XLSX kamu")
            bot.register_next_step_handler_by_chat_id(chat_id, handle_xlsx_to_vcf)
        elif text == "Hitung Kontak":
            bot.send_message(chat_id, "Silakan kirim file TXT/VCF/XLSX untuk dihitung")
        elif text == "Rename Kontak":
            bot.send_message(chat_id, "Fitur rename kontak masih placeholder")
        elif text == "Hapus Kontak":
            bot.send_message(chat_id, "Fitur hapus kontak masih placeholder")
        elif text == "Pecah File":
            bot.send_message(chat_id, "Fitur pecah file masih placeholder")
        elif text == "Gabung File":
            bot.send_message(chat_id, "Fitur gabung file masih placeholder")
        else:
            bot.send_message(chat_id, "Menu tidak dikenal")
    except Exception as e:
        bot.send_message(chat_id, f"Terjadi error: {e}")

# ===== FILE HANDLER PLACEHOLDER =====
def handle_txt_to_vcf(message):
    bot.send_message(message.chat.id, "TXT → VCF masih placeholder")

def handle_vcf_to_txt(message):
    bot.send_message(message.chat.id, "VCF → TXT masih placeholder")

def handle_xlsx_to_vcf(message):
    bot.send_message(message.chat.id, "XLSX → VCF masih placeholder")

# ===== KEEP BOT RUNNING =====
while True:
    try:
        bot.infinity_polling()
    except Exception as e:
        print("Error:", e)
        time.sleep(5)
